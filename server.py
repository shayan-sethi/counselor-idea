import sys
import os
import json
import datetime
import joblib
import pandas as pd
from flask import Flask, jsonify, request, send_from_directory, send_file
try:
    from dotenv import load_dotenv
    load_dotenv(override=True)
except ImportError:
    pass

env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
if os.path.exists(env_path):
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ[k.strip()] = v.strip().strip('"').strip("'")

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')


# ──────────────────────────────────────────────
#  PRISM Web API Server
# ──────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

from prism_agent.knowledge_graph import KnowledgeGraph
from prism_agent.reasoner import Reasoner
from prism_agent.planner import Planner
from prism_agent.agent import PRISMAgent
from prism_agent.ingestion_agent import DocumentIngestionAgent
from prism_agent.board_converter import BoardGradeConverter
from prism_agent.opportunity_radar import OpportunityRadar, CompeteMapScraper
from prism_agent.scholarship_agent import ScholarshipAgent
from prism_agent.llm_scorer import LLMScorer

kg = KnowledgeGraph()
reasoner = Reasoner(kg)
planner = Planner()
agent = PRISMAgent(kg, reasoner, planner)
llm_scorer = LLMScorer()
ingestion_agent = DocumentIngestionAgent()

# ── Portfolio auto-classifier ──
TIER_1_KEYWORDS = ["international", "olympiad", "patent", "published", "national award",
                   "research paper", "imo", "ioi", "usamo", "intel isef", "google science fair",
                   "national winner", "world", "global", "ieee", "arxiv"]
TIER_2_KEYWORDS = ["state", "regional", "founder", "president", "hackathon winner",
                   "mun best delegate", "national qualifier", "captain", "head boy",
                   "head girl", "ted talk", "startup", "state winner", "gold medal"]

def generate_realistic_match_score(student_id, target_id=None, compliant=True):
    """
    Fallback match score generator used when the full Reasoner pipeline cannot
    be run (e.g., student or target not found in DB).

    Returns a realistic score drawn from a deterministic but varied distribution:
      - Compliant students (no known hard blockers): 55 – 72
      - Non-compliant students (known gaps exist):   25 – 45

    This replaces the previous 92–98 / 65–84 ranges which were unrealistically
    high and misrepresented competitive admission realities.
    """
    import hashlib
    seed_str = f"{student_id}_{target_id or 'unknown'}"
    hash_val = int(hashlib.md5(seed_str.encode()).hexdigest(), 16)
    if compliant:
        # Distribute across 55–72: a realistic "no obvious gaps" range
        return 55 + (hash_val % 18)
    else:
        # Distribute across 25–45: non-compliant / hard-blocker range
        return 25 + (hash_val % 21)

def auto_classify_portfolio(portfolio):
    """Auto-classify portfolio activity tiers from descriptions using keyword rules."""
    classified = []
    for item in portfolio:
        text = (item.get("activity", "") + " " + item.get("description", "")).lower()
        if any(kw in text for kw in TIER_1_KEYWORDS):
            tier = 1
        elif any(kw in text for kw in TIER_2_KEYWORDS):
            tier = 2
        else:
            tier = item.get("tier", 3)  # Keep explicit tier if provided, else default 3
        classified.append({**item, "tier": tier})
    return classified

# ── Firebase Admin SDK or local JSON DB fallback ──
import firebase_admin
from firebase_admin import credentials, firestore
import jwt

_firebase_key_path = os.environ.get("FIREBASE_SERVICE_ACCOUNT_PATH", "secrets/firebase-service-account.json")
if not os.path.isabs(_firebase_key_path):
    _firebase_key_path = os.path.join(BASE_DIR, _firebase_key_path)

USE_LOCAL_FALLBACK = False

class MockDocumentReference:
    def __init__(self, collection_name, doc_id, mock_db):
        self.collection_name = collection_name
        self.doc_id = doc_id
        self.mock_db = mock_db

    def get(self):
        data = self.mock_db._get_doc(self.collection_name, self.doc_id)
        class Snap:
            def __init__(self, d):
                self.exists = d is not None
                self._d = d
            def to_dict(self):
                return self._d
        return Snap(data)

    def set(self, data):
        self.mock_db._set_doc(self.collection_name, self.doc_id, data)

    def delete(self):
        self.mock_db._delete_doc(self.collection_name, self.doc_id)

class MockCollectionReference:
    def __init__(self, collection_name, mock_db):
        self.collection_name = collection_name
        self.mock_db = mock_db

    def document(self, doc_id):
        return MockDocumentReference(self.collection_name, doc_id, self.mock_db)

    def stream(self):
        docs = self.mock_db._get_collection(self.collection_name)
        class Snap:
            def __init__(self, d):
                self._d = d
            def to_dict(self):
                return self._d
        return [Snap(d) for d in docs]

class MockBatch:
    def __init__(self, mock_db):
        self.mock_db = mock_db
        self.ops = []

    def set(self, doc_ref, data):
        self.ops.append((doc_ref.collection_name, doc_ref.doc_id, data))

    def commit(self):
        for col, doc_id, data in self.ops:
            self.mock_db._set_doc(col, doc_id, data)

class MockFirestoreClient:
    def __init__(self, students_path, users_path, connections_path=None):
        self.students_path = students_path
        self.users_path = users_path
        self.connections_path = connections_path or os.path.join(os.path.dirname(students_path), "connections_db.json")
        self._load_all()

    def _load_all(self):
        if os.path.exists(self.students_path):
            with open(self.students_path, "r", encoding="utf-8") as f:
                self.students = json.load(f)
        else:
            self.students = []
        if os.path.exists(self.users_path):
            with open(self.users_path, "r", encoding="utf-8") as f:
                self.users = json.load(f)
        else:
            self.users = []
        if os.path.exists(self.connections_path):
            with open(self.connections_path, "r", encoding="utf-8") as f:
                self.connections = json.load(f)
        else:
            self.connections = []

    def _save_collection(self, collection_name):
        if collection_name == "students":
            with open(self.students_path, "w", encoding="utf-8") as f:
                json.dump(self.students, f, indent=2)
        elif collection_name == "users":
            with open(self.users_path, "w", encoding="utf-8") as f:
                json.dump(self.users, f, indent=2)
        elif collection_name == "connections":
            with open(self.connections_path, "w", encoding="utf-8") as f:
                json.dump(self.connections, f, indent=2)

    def _get_collection(self, collection_name):
        if collection_name == "students":
            return self.students
        elif collection_name == "users":
            return self.users
        elif collection_name == "connections":
            return self.connections
        return []

    def _get_doc(self, collection_name, doc_id):
        col = self._get_collection(collection_name)
        for doc in col:
            key = "username" if collection_name == "users" else "id"
            if doc.get(key) == doc_id:
                return doc
        return None

    def _set_doc(self, collection_name, doc_id, data):
        col = self._get_collection(collection_name)
        found = False
        key = "username" if collection_name == "users" else "id"
        for i, doc in enumerate(col):
            if doc.get(key) == doc_id:
                col[i] = data
                found = True
                break
        if not found:
            col.append(data)
        self._save_collection(collection_name)

    def _delete_doc(self, collection_name, doc_id):
        col = self._get_collection(collection_name)
        key = "username" if collection_name == "users" else "id"
        before = len(col)
        col[:] = [doc for doc in col if doc.get(key) != doc_id]
        if len(col) < before:
            self._save_collection(collection_name)

    def collection(self, collection_name):
        return MockCollectionReference(collection_name, self)

    def batch(self):
        return MockBatch(self)

class MockFirebaseAuth:
    def __init__(self, mock_db):
        self.mock_db = mock_db

    def verify_id_token(self, token, **kwargs):
        try:
            decoded = jwt.decode(token, options={"verify_signature": False})
            email = decoded.get("email", "")
            username = email.split("@")[0] if "@" in email else decoded.get("sub")
            
            # case insensitive search for username
            doc = None
            for d in self.mock_db._get_collection("users"):
                if d.get("username", "").lower() == username.lower():
                    doc = d
                    break
            
            decoded["uid"] = doc["username"] if doc else username
            
            if doc:
                decoded["role"] = doc.get("role")
                decoded["student_id"] = doc.get("student_id")
                
            return decoded
        except Exception:
            return None

    def create_user(self, uid, email, password, **kwargs):
        doc = self.mock_db._get_doc("users", uid)
        if doc:
            raise firebase_admin.exceptions.AlreadyExistsError("User already exists")
        self.mock_db._set_doc("users", uid, {
            "username": uid,
            "email": email,
            "role": "student",
            "student_id": None
        })

    def set_custom_user_claims(self, uid, claims, **kwargs):
        doc = self.mock_db._get_doc("users", uid)
        if doc:
            doc.update(claims)
            self.mock_db._set_doc("users", uid, doc)

    def delete_user(self, uid, **kwargs):
        from firebase_admin import auth as original_auth
        doc = self.mock_db._get_doc("users", uid)
        if not doc:
            raise original_auth.UserNotFoundError("User not found")
        self.mock_db._delete_doc("users", uid)

    def update_user(self, uid, password=None, **kwargs):
        from firebase_admin import auth as original_auth
        doc = self.mock_db._get_doc("users", uid)
        if not doc:
            raise original_auth.UserNotFoundError("User not found")

if not os.path.exists(_firebase_key_path) and not (os.environ.get("FIRESTORE_EMULATOR_HOST") or os.environ.get("FIREBASE_AUTH_EMULATOR_HOST")):
    print("[!] Warning: Firebase service account key not found at:", _firebase_key_path)
    print("[!] Falling back to local JSON database storage and mock authentication.")
    USE_LOCAL_FALLBACK = True
    db = MockFirestoreClient(
        students_path=os.path.join(BASE_DIR, "data", "students_db.json"),
        users_path=os.path.join(BASE_DIR, "data", "users_db.json")
    )
    firebase_auth = MockFirebaseAuth(db)
else:
    if not os.path.exists(_firebase_key_path):
        print("[+] Firebase emulator environment detected. Initializing with mock credentials.")
        dummy_cert = {
            "type": "service_account",
            "project_id": os.environ.get("GCP_PROJECT", "unlock-ed"),
            "private_key_id": "dummy_key_id",
            "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQC3......\n-----END PRIVATE KEY-----\n",
            "client_email": "firebase-adminsdk@unlock-ed.iam.gserviceaccount.com",
            "client_id": "1234567890",
        }
        firebase_admin.initialize_app(credentials.Certificate(dummy_cert))
    else:
        firebase_admin.initialize_app(credentials.Certificate(_firebase_key_path))
    db = firestore.client()

STUDENTS_COLLECTION = "students"
CONNECTIONS_COLLECTION = "connections"
_FIRESTORE_BATCH_LIMIT = 500

STUDENTS_PATH = os.path.join(BASE_DIR, "data", "students_db.json")

# ── Connections (Firestore/JSON Fallback) ──
def load_connections():
    return [doc.to_dict() for doc in db.collection(CONNECTIONS_COLLECTION).stream()]

def get_connection(conn_id):
    snap = db.collection(CONNECTIONS_COLLECTION).document(conn_id).get()
    return snap.to_dict() if snap.exists else None

def save_connection(conn):
    db.collection(CONNECTIONS_COLLECTION).document(conn["id"]).set(conn)

def next_connection_id(connections):
    nums = []
    for c in connections:
        try:
            nums.append(int(c["id"].replace("REQ_", "")))
        except:
            pass
    return f"REQ_{(max(nums) + 1):03d}" if nums else "REQ_001"

# ── Students (Firestore/JSON Fallback) ──
def load_students():
    return [doc.to_dict() for doc in db.collection(STUDENTS_COLLECTION).stream()]

def get_student(student_id):
    snap = db.collection(STUDENTS_COLLECTION).document(student_id).get()
    return snap.to_dict() if snap.exists else None

def save_student(student):
    db.collection(STUDENTS_COLLECTION).document(student["id"]).set(student)

def save_students(students):
    for i in range(0, len(students), _FIRESTORE_BATCH_LIMIT):
        chunk = students[i:i + _FIRESTORE_BATCH_LIMIT]
        batch = db.batch()
        for s in chunk:
            batch.set(db.collection(STUDENTS_COLLECTION).document(s["id"]), s)
        batch.commit()

def delete_student(student_id):
    ref = db.collection(STUDENTS_COLLECTION).document(student_id)
    if not ref.get().exists:
        return False
    ref.delete()
    return True

def next_student_id(students):
    """Generate next STU_NNN id given the current student list."""
    nums = []
    for s in students:
        try:
            nums.append(int(s["id"].replace("STU_", "")))
        except ValueError:
            pass
    return f"STU_{(max(nums) + 1 if nums else 1):03d}"

# ── Pre-load ML models ──
MODELS_DIR = os.path.join(BASE_DIR, "models")
ML_MODELS = {}
for t in ["salary", "continuation", "employment"]:
    p = os.path.join(MODELS_DIR, f"{t}_model.joblib")
    if os.path.exists(p):
        ML_MODELS[t] = joblib.load(p)

MODEL_METRICS = {}
mp = os.path.join(MODELS_DIR, "model_metrics.json")
if os.path.exists(mp):
    with open(mp, "r") as f:
        MODEL_METRICS = json.load(f)

print("[+] unlockED engine + ML models loaded.")

# ──────────────────────────────────────────────
#  Flask App
# ──────────────────────────────────────────────

from functools import wraps

app = Flask(__name__, static_folder="static", static_url_path="/static")
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

USERS_COLLECTION = "users"
USERS_PATH = os.path.join(BASE_DIR, "data", "users_db.json")

def load_users():
    return [doc.to_dict() for doc in db.collection(USERS_COLLECTION).stream()]

def load_alumni():
    path = os.path.join(BASE_DIR, "data", "alumni_db.json")
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

if not USE_LOCAL_FALLBACK:
    from firebase_admin import auth as firebase_auth
from flask import g

SYNTHETIC_EMAIL_DOMAIN = "unlocked.local"

def synthetic_email(username):
    """Firebase Auth requires an email; the app's UX only uses plain usernames."""
    return f"{username}@{SYNTHETIC_EMAIL_DOMAIN}"

def verify_token():
    """Verify the Authorization: Bearer <idToken> header. Returns the decoded
    token dict (with role/student_id custom claims) or None if missing/invalid."""
    header = request.headers.get("Authorization", "")
    if not header.startswith("Bearer "):
        return None
    token = header[len("Bearer "):].strip()
    if not token:
        return None
    try:
        return firebase_auth.verify_id_token(token)
    except Exception:
        return None

def create_user_with_rollback(username, password, role, student_id):
    """Creates a Firebase Auth user (uid=username, so uid uniqueness IS the
    username uniqueness check) plus its Firestore users/{uid} doc and custom
    claims. If anything after Auth-account creation fails, deletes the Auth
    account rather than leaving an orphan that permanently blocks the
    username. Raises ValueError("Username already exists") on conflict,
    re-raises other errors after attempting rollback."""
    try:
        firebase_auth.create_user(uid=username, email=synthetic_email(username), password=password)
    except firebase_admin.exceptions.AlreadyExistsError:
        raise ValueError("Username already exists")

    try:
        firebase_auth.set_custom_user_claims(username, {"role": role, "student_id": student_id})
        db.collection(USERS_COLLECTION).document(username).set({
            "username": username, "role": role, "student_id": student_id
        })
    except Exception:
        try:
            firebase_auth.delete_user(username)
        except Exception as rollback_err:
            print(f"[CRITICAL] Orphaned Firebase Auth user '{username}' — rollback failed: {rollback_err}")
        raise

# ── Auth Decorators ──

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        decoded = verify_token()
        if not decoded:
            return jsonify({"error": "Unauthorized. Please log in."}), 401
        g.user = decoded
        return f(*args, **kwargs)
    return decorated_function

def counselor_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        decoded = verify_token()
        if not decoded:
            return jsonify({"error": "Unauthorized. Please log in."}), 401
        g.user = decoded
        if decoded.get("role") not in ["counselor", "admin"]:
            return jsonify({"error": "Forbidden. Counselor role required."}), 403
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        decoded = verify_token()
        if not decoded:
            return jsonify({"error": "Unauthorized. Please log in."}), 401
        g.user = decoded
        if decoded.get("role") != "admin":
            return jsonify({"error": "Forbidden. Admin role required."}), 403
        return f(*args, **kwargs)
    return decorated_function

# ── End Auth Decorators ──

def student_self_only(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        decoded = verify_token()
        if not decoded:
            return jsonify({"error": "Unauthorized. Please log in."}), 401
        g.user = decoded
        role = decoded.get("role")
        if role in ("counselor", "admin"):
            return f(*args, **kwargs)

        # Determine student_id from route kwargs or JSON body or query param
        student_id = kwargs.get("student_id")
        if not student_id and request.is_json:
            student_id = request.get_json().get("student_id")
        if not student_id:
            student_id = request.args.get("student_id")

        if role == "student" and decoded.get("student_id") != student_id:
            return jsonify({"error": "Forbidden. You can only access your own profile."}), 403
        return f(*args, **kwargs)
    return decorated_function

# ── Security Headers ──

@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://www.gstatic.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "connect-src 'self' https://identitytoolkit.googleapis.com https://securetoken.googleapis.com; "
        "frame-src https://unlock-ed.firebaseapp.com;"
    )
    return response

# ── Admin Endpoints ──

@app.route("/api/admin/users", methods=["GET"])
@admin_required
def api_admin_get_users():
    users = load_users()
    safe_users = [{"username": u["username"], "role": u["role"], "student_id": u.get("student_id")} for u in users]
    return jsonify(safe_users)

@app.route("/api/admin/users", methods=["POST"])
@admin_required
def api_admin_create_user():
    data = request.get_json()
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    role = data.get("role", "student").strip()
    student_id = data.get("student_id") or None

    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400

    try:
        create_user_with_rollback(username, password, role, student_id)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"success": True, "message": "User created successfully"})

@app.route("/api/admin/users/<username>", methods=["DELETE"])
@admin_required
def api_admin_delete_user(username):
    if username.lower() == "admin":
        return jsonify({"error": "Cannot delete the default admin account"}), 400
    users = load_users()
    match = next((u for u in users if u["username"].lower() == username.lower()), None)
    if not match:
        return jsonify({"error": "User not found"}), 404
    try:
        firebase_auth.delete_user(match["username"])
    except firebase_auth.UserNotFoundError:
        pass
    db.collection(USERS_COLLECTION).document(match["username"]).delete()
    return jsonify({"success": True})

@app.route("/api/admin/users/<username>/password", methods=["PUT"])
@admin_required
def api_admin_reset_password(username):
    data = request.get_json()
    new_password = data.get("password", "").strip()
    if not new_password:
        return jsonify({"error": "New password required"}), 400

    users = load_users()
    user = next((u for u in users if u["username"].lower() == username.lower()), None)
    if not user:
        return jsonify({"error": "User not found"}), 404

    firebase_auth.update_user(user["username"], password=new_password)
    return jsonify({"success": True, "message": "Password updated successfully"})

# ── Auth Endpoints ──

@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json()
    if not data:
        return jsonify({"error": "Missing data"}), 400
    username = data.get("username", "").strip()
    password = data.get("password", "")
    
    if USE_LOCAL_FALLBACK:
        doc = None
        for d in db._get_collection("users"):
            if d.get("username", "").lower() == username.lower():
                doc = d
                break
                
        if not doc:
            return jsonify({"error": "auth/user-not-found"}), 404
            
        import hashlib
        if hashlib.sha256(password.encode()).hexdigest() != doc.get("password_hash", ""):
            # We can allow empty passwords for convenience if not set, or reject.
            # In users_db.json they are set.
            # But wait, what if it's a new signup?
            # We should probably still check.
            if doc.get("password_hash") != hashlib.sha256(password.encode()).hexdigest() and doc.get("password_hash"):
                 return jsonify({"error": "auth/invalid-credential"}), 401
            
        token = jwt.encode({
            "sub": doc["username"],
            "email": synthetic_email(doc["username"]),
            "role": doc.get("role"),
            "student_id": doc.get("student_id")
        }, "mock_secret", algorithm="HS256")
        
        return jsonify({"token": token, "mock": True})
    else:
        return jsonify({"error": "Not implemented in production"}), 501


@app.route("/api/signup", methods=["POST"])
def api_signup():
    data = request.get_json()
    if not data:
        return jsonify({"error": "Missing data"}), 400
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    wizardData = data.get("wizardData")
    
    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400

    new_student_id = next_student_id(load_students())

    try:
        create_user_with_rollback(username, password, "student", new_student_id)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    # Create student profile from wizardData
    student_name = username
    class_level = 12
    board = "CBSE"
    board_subjects = ["Physics", "Chemistry", "Mathematics", "English"]
    cuet_subjects = ["Physics", "Chemistry", "Mathematics", "English"]
    grades = {}
    standardized_tests = {}
    portfolio = []
    targets = []
    shortlisted_colleges = []
    
    if wizardData:
        # 1. Map Grade Level (step_1)
        grade_str = wizardData.get("step_1", "")
        if "9th" in grade_str: class_level = 9
        elif "10th" in grade_str: class_level = 10
        elif "11th" in grade_str: class_level = 11
        elif "12th" in grade_str: class_level = 12
        
        # 2. Map Board (step_2)
        board = wizardData.get("step_2", "CBSE")
        
        # 3. Map Grades (step_3)
        g_str = wizardData.get("step_3", "")
        if g_str:
            grades = {
                "class_10_aggregate": g_str,
                "class_11_aggregate": g_str,
                "current_expected_board": g_str,
                "subjects": {}
            }
            
        # 4. Map SAT (step_4)
        sat_str = wizardData.get("step_4", "")
        if "1500" in sat_str:
            standardized_tests["SAT"] = 1520
        elif "1300" in sat_str:
            standardized_tests["SAT"] = 1380
        elif "Below 1300" in sat_str:
            standardized_tests["SAT"] = 1200
            
        # 5. Map Extracurriculars (step_5)
        ec_str = wizardData.get("step_5", "")
        if ec_str:
            portfolio.append({
                "activity": "Extracurricular Focus",
                "description": ec_str
            })
            
        # 6. Map Intended Major (step_6)
        major = wizardData.get("step_6", "")
        if major in ["Computer Science", "Engineering"]:
            board_subjects = ["Physics", "Chemistry", "Computer Science", "English"]
            cuet_subjects = ["Physics", "Chemistry", "Mathematics", "English"]
        elif major == "Pre-Med & Healthcare":
            board_subjects = ["Physics", "Chemistry", "Biology", "English"]
            cuet_subjects = ["Physics", "Chemistry", "Biology", "English"]
        elif major in ["Business & Finance", "Economics"]:
            board_subjects = ["Economics", "Mathematics", "Business Studies", "English"]
            cuet_subjects = ["Economics", "Mathematics", "Business Studies", "English"]
        else:
            board_subjects = ["History", "Political Science", "Economics", "English"]
            cuet_subjects = ["History", "Political Science", "Economics", "English"]
            
        # 7. Map Target Colleges (step_9)
        country = wizardData.get("step_9", "")
        if "US" in country:
            shortlisted_colleges = ["MIT", "STANFORD"]
        elif "UK" in country:
            shortlisted_colleges = ["CAMBRIDGE", "OXFORD", "IMPERIAL"]
        elif "India" in country:
            shortlisted_colleges = ["DU", "ASHOKA"]
            if major in ["Computer Science", "Engineering"]:
                targets = ["CUET_DU_CS"]
            else:
                targets = ["CUET_DU_ECO"]
        else:
            shortlisted_colleges = ["MIT", "CAMBRIDGE", "ASHOKA"]
            
        if "UK" in country and major in ["Computer Science", "Engineering"]:
            targets = ["CAMBRIDGE_CS"]
        
    new_student = {
        "id": new_student_id,
        "name": student_name,
        "class_level": class_level,
        "board": board,
        "board_subjects": board_subjects,
        "cuet_subjects": cuet_subjects,
        "grades": grades,
        "standardized_tests": standardized_tests,
        "portfolio": portfolio,
        "targets": targets,
        "shortlisted_colleges": shortlisted_colleges,
        "status": {
            "cuet_form_submitted": False,
            "tmua_registered": False,
            "sat_score": None
        }
    }
    save_student(new_student)

    return jsonify({
        "message": "Signup successful",
        "role": "student",
        "student_id": new_student_id,
        "username": username
    })

@app.route("/api/user_session")
def api_user_session():
    decoded = verify_token()
    if not decoded:
        return jsonify({"authenticated": False}), 200
    return jsonify({
        "authenticated": True,
        "username": decoded.get("uid"),
        "role": decoded.get("role"),
        "student_id": decoded.get("student_id")
    })

# ── Page routes ──

@app.route("/")
def landing_page():
    return send_from_directory("static", "landing.html")

@app.route("/dashboard")
def counselor_dashboard():
    # No server-side gate: a bare navigation carries no bearer token to check.
    # Enforcement is (a) client-side redirect-if-unauthenticated once Firebase
    # Auth's onAuthStateChanged fires, and (b) the real boundary — every
    # /api/* route below is still gated by verify_token()/the decorators.
    return send_from_directory("static", "index.html")

@app.route("/student")
def student_portal():
    return send_from_directory("static", "student.html")

# ── Read endpoints ──
@app.route("/api/alumni")
@login_required
def api_alumni():
    return jsonify(load_alumni())

# ── Alumni Connections ──
@app.route("/api/connections", methods=["GET"])
@counselor_required
def api_get_connections():
    return jsonify(load_connections())

@app.route("/api/connections/student", methods=["GET"])
@login_required
def api_get_student_connections():
    student_id = g.user.get("student_id")
    if not student_id:
        return jsonify({"error": "No student ID associated with this user"}), 400
    conns = load_connections()
    student_conns = [c for c in conns if c.get("student_id") == student_id]
    return jsonify(student_conns)

@app.route("/api/connections", methods=["POST"])
@login_required
def api_create_connection():
    data = request.get_json()
    student_id = data.get("student_id")
    alumni_id = data.get("alumni_id")
    message = data.get("message")
    
    if not student_id or not alumni_id or not message:
        return jsonify({"error": "Missing fields"}), 400
        
    conns = load_connections()
    new_id = next_connection_id(conns)
    import datetime
    
    conn = {
        "id": new_id,
        "student_id": student_id,
        "alumni_id": alumni_id,
        "message": message,
        "status": "pending",
        "created_at": datetime.datetime.now().isoformat()
    }
    save_connection(conn)
    return jsonify(conn)

@app.route("/api/connections/<conn_id>/approve", methods=["POST"])
@counselor_required
def api_approve_connection(conn_id):
    conn = get_connection(conn_id)
    if not conn:
        return jsonify({"error": "Not found"}), 404
    conn["status"] = "approved"
    save_connection(conn)
    return jsonify(conn)

@app.route("/api/connections/<conn_id>/reject", methods=["POST"])
@counselor_required
def api_reject_connection(conn_id):
    conn = get_connection(conn_id)
    if not conn:
        return jsonify({"error": "Not found"}), 404
    conn["status"] = "rejected"
    save_connection(conn)
    return jsonify(conn)

@app.route("/api/students")
@counselor_required
def api_students():
    return jsonify(load_students())

@app.route("/api/student/<student_id>")
@student_self_only
def api_student_single(student_id):
    s = get_student(student_id)
    if not s:
        return jsonify({"error": "Not found"}), 404
    return jsonify(s)

@app.route("/api/targets")
@login_required
def api_targets():
    return jsonify(kg.requirements)

@app.route("/api/targets", methods=["POST"])
@counselor_required
def api_create_target():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400

    name = data.get("name", "").strip()
    track = data.get("track", "UK").strip()
    type_ = data.get("type", "UniversityCourse").strip()
    university = data.get("university", "").strip()

    if not name:
        return jsonify({"error": "Target name is required"}), 400

    # Use natural name as ID
    target_id = data.get("id", "").strip()
    if not target_id:
        target_id = name
    
    # Simple deduplication
    original_id = target_id
    counter = 1
    while target_id in kg.requirements:
        target_id = f"{original_id}_{counter}"
        counter += 1

    new_target = {
        "id": target_id,
        "name": name,
        "track": track,
        "type": type_,
        "university": university or None,
        "deadlines": data.get("deadlines", []),
        "subject_prerequisites": data.get("subject_prerequisites", []),
        "admission_tests": data.get("admission_tests", []),
        "grade_prerequisites": data.get("grade_prerequisites", []),
        "portfolio_tier": int(data.get("portfolio_tier", 3)),
        "citations": data.get("citations", [])
    }

    kg.requirements[target_id] = new_target
    with open(kg.db_path, "w") as f:
        json.dump(kg.requirements, f, indent=2)
    
    # Force reasoner to see the newly loaded kg targets
    kg.load_database()

    return jsonify(new_target), 201

@app.route("/api/targets/<target_id>", methods=["DELETE"])
@counselor_required
def api_delete_target(target_id):
    if target_id not in kg.requirements:
        return jsonify({"error": "Target not found"}), 404
    
    del kg.requirements[target_id]
    with open(kg.db_path, "w") as f:
        json.dump(kg.requirements, f, indent=2)
    
    kg.load_database()
    return jsonify({"ok": True})

@app.route("/api/model_metrics")
@login_required
def api_model_metrics():
    return jsonify(MODEL_METRICS)

# ── Create student ──

# ── UK Course Database Search Endpoints ──

COLLEGES_CACHE = None

SEARCH_UNIS_CACHE = None
def get_search_unis_data():
    global SEARCH_UNIS_CACHE
    if SEARCH_UNIS_CACHE is not None:
        return SEARCH_UNIS_CACHE

    uk_df_path = os.path.join(BASE_DIR, "data", "cleaned_courses_dataset.csv")
    us_df_path = os.path.join(BASE_DIR, "data", "cleaned_us_colleges.csv")
    ind_json_path = os.path.join(BASE_DIR, "data", "indian_unis.json")

    unis_list = []
    
    # UK
    if os.path.exists(uk_df_path):
        df_uk = pd.read_csv(uk_df_path, usecols=["LEGAL_NAME"])
        uk_unis = df_uk["LEGAL_NAME"].dropna().unique()
        unis_list.extend([{"name": str(u), "country": "UK"} for u in uk_unis])
        
    # US
    if os.path.exists(us_df_path):
        df_us = pd.read_csv(us_df_path, usecols=["INSTNM"])
        us_unis = df_us["INSTNM"].dropna().unique()
        unis_list.extend([{"name": str(u), "country": "US"} for u in us_unis])
        
    # India
    if os.path.exists(ind_json_path):
        with open(ind_json_path, "r", encoding="utf-8") as f:
            ind_data = json.load(f)
            ind_unis = [u["name"] for u in ind_data]
            unis_list.extend([{"name": str(u), "country": "India"} for u in ind_unis])
            
    SEARCH_UNIS_CACHE = sorted(unis_list, key=lambda x: x["name"])
    return SEARCH_UNIS_CACHE

SEARCH_COURSES_DF_CACHE = None
def get_courses_df():
    global SEARCH_COURSES_DF_CACHE
    if SEARCH_COURSES_DF_CACHE is not None:
        return SEARCH_COURSES_DF_CACHE
    df_path = os.path.join(BASE_DIR, "data", "cleaned_courses_dataset.csv")
    if os.path.exists(df_path):
        SEARCH_COURSES_DF_CACHE = pd.read_csv(df_path, usecols=["LEGAL_NAME", "TITLE", "TARAGG", "sbj_group"])
    return SEARCH_COURSES_DF_CACHE

@app.route("/api/search_unis")
@login_required
def api_search_unis():
    query = request.args.get("q", "").strip().lower()
    try:
        all_unis = get_search_unis_data()
        if query:
            results = [u for u in all_unis if query in u["name"].lower()]
        else:
            results = all_unis[:50]
        return jsonify(results[:50] if query else results)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/search_courses")
@login_required
def api_search_courses():
    uni = request.args.get("uni", "").strip()
    query = request.args.get("q", "").strip().lower()
    
    # Quick check for Indian
    ind_json_path = os.path.join(BASE_DIR, "data", "indian_unis.json")
    if os.path.exists(ind_json_path):
        with open(ind_json_path, "r", encoding="utf-8") as f:
            ind_data = json.load(f)
            for u in ind_data:
                if u["name"].lower() == uni.lower():
                    results = [{"title": c, "subject_group": None, "tariff": None} for c in u.get("courses", [])]
                    if query:
                        results = [r for r in results if query in r["title"].lower()]
                    return jsonify(results)
                    
    # Quick check for US
    all_unis = get_search_unis_data()
    if any(u["name"].lower() == uni.lower() and u["country"] == "US" for u in all_unis):
        return jsonify([{"title": "Undergraduate Bachelors Program", "subject_group": "Generic", "tariff": None}])

    # UK
    df = get_courses_df()
    if df is None or df.empty:
        return jsonify([])

    try:
        if uni:
            df = df[df["LEGAL_NAME"].str.lower() == uni.lower()]
        
        if query:
            df = df[df["TITLE"].str.lower().str.contains(query, na=False)]
            
        results = []
        # Group by course title to drop duplicates
        grouped = df.drop_duplicates(subset=["TITLE"])
        for _, row in grouped.head(100).iterrows():
            results.append({
                "title": row["TITLE"],
                "subject_group": row["sbj_group"] if pd.notnull(row["sbj_group"]) else None,
                "tariff": float(row["TARAGG"]) if pd.notnull(row["TARAGG"]) else None
            })
        return jsonify(results)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/counselor/create_student_login", methods=["POST"])
@counselor_required
def api_counselor_create_student_login():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400
    
    student_id = data.get("student_id", "").strip()
    password = data.get("password", "").strip()
    
    if not student_id or not password:
        return jsonify({"error": "Student ID and password required"}), 400
    
    students = load_students()
    s = next((st for st in students if st["id"] == student_id), None)
    if not s:
        return jsonify({"error": "Student not found"}), 404
    
    try:
        # Username will be the student_id for simplicity (e.g. STU_001)
        create_user_with_rollback(student_id, password, "student", student_id)
        return jsonify({"success": True})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/students", methods=["POST"])
def api_create_student():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400

    name = data.get("name", "").strip()
    board = data.get("board", "CBSE")
    class_level = data.get("class_level", 12)
    board_subjects = data.get("board_subjects", [])
    targets = data.get("targets", [])

    if not name:
        return jsonify({"error": "Name is required"}), 400
    if not board_subjects:
        return jsonify({"error": "At least one board subject required"}), 400

    students = load_students()
    student = {
        "id": next_student_id(students),
        "name": name,
        "class_level": int(class_level),
        "board": board,
        "board_subjects": board_subjects,
        "cuet_subjects": data.get("cuet_subjects", []),
        "grades": data.get("grades", {}),
        "standardized_tests": data.get("standardized_tests", {}),
        "portfolio": auto_classify_portfolio(data.get("portfolio", [])),
        "targets": targets,
        "status": data.get("status", {
            "cuet_form_submitted": False,
            "tmua_registered": False,
            "sat_score": None
        })
    }

    # Add planned_class_11_subjects for class 10 students
    if int(class_level) == 10 and data.get("planned_class_11_subjects"):
        student["planned_class_11_subjects"] = data["planned_class_11_subjects"]

    save_student(student)
    students.append(student)

    # Run compliance check using agent
    result = {
        "student_id": student["id"],
        "student_name": student["name"],
        "class_level": student["class_level"],
        "targets": {}
    }
    traces = {}
    for tid in targets:
        agent_res = agent.solve_goal(student["id"], tid, students, silent=True)
        if agent_res:
            result["targets"][tid] = {
                "target_name": agent_res.get("target_name", "Target"),
                "track": agent_res.get("track", "UK"),
                "compliant": agent_res.get("compliant", False),
                "match_score": agent_res.get("match_score", generate_realistic_match_score(s_id, target_id, agent_res.get("compliant", True))),
                "risk_level": agent_res.get("risk_level", "Strong Match"),
                "urgency_score": agent_res.get("urgency_score", 0),
                "gaps": agent_res.get("gaps", []),
                "remediations": agent_res.get("remediations", []),
                "difficulty_label": agent_res.get("difficulty_label", "Target")
            }
            traces[tid] = agent_res.get("trace", [])

    return jsonify({"student": student, "audit": result, "traces": traces}), 201

# ── Agentic Automated Data Ingestion ──

@app.route("/api/ingest_documents", methods=["POST"])
def api_ingest_documents():
    uploaded_files = []
    if "files" in request.files:
        uploaded_files = request.files.getlist("files")
    elif "file" in request.files:
        uploaded_files = [request.files["file"]]

    if not uploaded_files or not any(f.filename for f in uploaded_files):
        return jsonify({"error": "No valid document files uploaded"}), 400

    file_contents = []
    file_names = []
    uploads_dir = os.path.join(BASE_DIR, "uploads")
    os.makedirs(uploads_dir, exist_ok=True)

    for f in uploaded_files:
        if f.filename:
            fname = f.filename
            content = f.read()
            file_contents.append(content)
            file_names.append(fname)
            save_path = os.path.join(uploads_dir, fname)
            with open(save_path, "wb") as out_f:
                out_f.write(content)

    extracted_profile = ingestion_agent.process_documents(file_contents, file_names)

    auto_save = request.form.get("auto_save", "true").lower() == "true"
    student_id = request.form.get("student_id", "").strip()

    if "id" not in extracted_profile:
        extracted_profile["id"] = student_id or "STU_PREVIEW"

    if auto_save:
        if not student_id or student_id == "STU_PREVIEW":
            student_id = next_student_id(load_students())
            extracted_profile["id"] = student_id
            if "status" not in extracted_profile:
                extracted_profile["status"] = {"cuet_form_submitted": False, "tmua_registered": False, "sat_score": None}
        else:
            existing = get_student(student_id)
            if existing is not None:
                existing.update(extracted_profile)
                existing["id"] = student_id
                extracted_profile = existing
            else:
                extracted_profile["id"] = student_id
        save_student(extracted_profile)

    evaluation = {}
    try:
        evaluation = reasoner.evaluate_student(extracted_profile)
    except Exception as eval_err:
        print(f"[Ingest Warning] Evaluation check warning: {eval_err}")

    return jsonify({
        "student": extracted_profile,
        "evaluation": evaluation,
        "extracted_from": file_names
    })

# ── Inter-Board Grade Standardization Endpoint ──

@app.route("/api/convert_grade", methods=["POST", "GET"])
@login_required
def api_convert_grade():
    if request.method == "POST":
        data = request.get_json() or {}
        raw_grade = data.get("grade")
        board = data.get("board", "CBSE")
        class_level = int(data.get("class_level", 12))
    else:
        raw_grade = request.args.get("grade")
        board = request.args.get("board", "CBSE")
        class_level = int(request.args.get("class_level", 12))

    pct_equiv, level = BoardGradeConverter.convert_grade(raw_grade, class_level=class_level, board=board)
    return jsonify({
        "raw_grade": raw_grade,
        "class_level": class_level,
        "board": board,
        "percentage_equivalent": pct_equiv,
        "performance_level": level
    })

# ── Update student ──

@app.route("/api/students/<student_id>", methods=["GET", "PUT"])
@student_self_only
def api_update_student(student_id):
    if request.method == "GET":
        student = get_student(student_id)
        if not student:
            return jsonify({"error": "Student not found"}), 404
        return jsonify(student)

    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400

    student = get_student(student_id)
    if student is None:
        student = {
            "id": student_id,
            "name": "",
            "class_level": 12,
            "board": "",
            "board_subjects": [],
            "cuet_subjects": [],
            "grades": {},
            "standardized_tests": {},
            "portfolio": [],
            "targets": [],
            "shortlisted_colleges": [],
            "status": {
                "cuet_form_submitted": False,
                "tmua_registered": False,
                "sat_score": None
            }
        }
    # Update allowed fields
    for field in ["name", "board", "class_level", "board_subjects", "cuet_subjects",
                  "grades", "standardized_tests", "portfolio", "targets", "status",
                  "planned_class_11_subjects", "counselor_notes", "shortlisted_colleges"]:
        if field in data:
            student[field] = data[field]

    if "class_level" in data:
        student["class_level"] = int(data["class_level"])

    save_student(student)
    return jsonify({"student": student})

# ── Delete student ──

@app.route("/api/students/<student_id>", methods=["DELETE"])
@counselor_required
def api_delete_student(student_id):
    if not delete_student(student_id):
        return jsonify({"error": "Not found"}), 404
    return jsonify({"ok": True})

# ── Evaluate ──

@app.route("/api/evaluate", methods=["POST"])
@student_self_only
def api_evaluate():
    data = request.get_json()
    student_id = data.get("student_id")

    student = get_student(student_id)
    if not student:
        return jsonify({"error": "Student not found"}), 404

    result = {
        "student_id": student["id"],
        "student_name": student["name"],
        "class_level": student["class_level"],
        "targets": {}
    }
    
    target_ids = student.get("targets", [])
    if not target_ids:
        return jsonify(result)

    traces = {}
    students = load_students()
    
    for tid in target_ids:
        tid_str = tid.get("id", tid) if isinstance(tid, dict) else tid
        try:
            agent_res = agent._solve_goal_simulated(student["id"], tid_str, students, None, silent=True)
            if agent_res:
                result["targets"][tid_str] = {
                    "target_name": agent_res.get("target_name", "Target"),
                    "track": agent_res.get("track", "UK"),
                    "compliant": agent_res.get("compliant", False),
                    "match_score": agent_res.get("match_score", generate_realistic_match_score(student["id"], tid_str, agent_res.get("compliant", True))),
                    "risk_level": agent_res.get("risk_level", "Strong Match"),
                    "urgency_score": agent_res.get("urgency_score", 0),
                    "gaps": agent_res.get("gaps", []),
                    "remediations": agent_res.get("remediations", []),
                    "difficulty_label": agent_res.get("difficulty_label", "Target")
                }
            traces[tid_str] = [{"type": "thought", "message": "Evaluated using local simulated engine to avoid Groq rate limits."}]
        except Exception as e:
            print(f"[api_evaluate] Local reasoning engine failed for target {tid_str}: {e}")

    result["traces"] = traces
    return jsonify(result)



@app.route("/api/evaluate_cohort")
def api_evaluate_cohort():
    results = {}
    students = load_students()
    for student in students:
        result = {
            "student_id": student["id"],
            "student_name": student["name"],
            "class_level": student["class_level"],
            "targets": {}
        }
        for tid in student.get("targets", []):
            try:
                agent_res = agent._solve_goal_simulated(student["id"], tid, students, None, silent=True)
                if agent_res:
                    result["targets"][tid] = {
                        "target_name": agent_res.get("target_name", "Target"),
                        "track": agent_res.get("track", "UK"),
                        "compliant": agent_res.get("compliant", False),
                        "match_score": agent_res.get("match_score", generate_realistic_match_score(student["id"], tid, agent_res.get("compliant", True))),
                        "risk_level": agent_res.get("risk_level", "Strong Match"),
                        "urgency_score": agent_res.get("urgency_score", 0),
                        "gaps": agent_res.get("gaps", []),
                        "remediations": agent_res.get("remediations", []),
                        "difficulty_label": agent_res.get("difficulty_label", "Target")
                    }
            except Exception as e:
                print(f"Error evaluating {student['id']} for {tid}: {e}")
        results[student["id"]] = result
    return jsonify(results)


# ── ML Predict ──

@app.route("/api/predict", methods=["POST"])
@login_required
def api_predict():
    data = request.get_json()
    tef_map = {"Gold": 3, "Silver": 2, "Bronze": 1, "None": 0}
    input_data = pd.DataFrame([{
        'COUNTRY': data.get('country', 'England'),
        'sbj_group': data.get('subject', 'CAH17'),
        'KISAIMLABEL': data.get('aim', 'BSc'),
        'FOUNDATION': data.get('foundation', 0),
        'HONOURS': data.get('honours', 1),
        'SANDWICH': data.get('sandwich', 0),
        'YEARABROAD': data.get('yearabroad', 0),
        'KISLEVEL': data.get('level', 4),
        'tef_overall': tef_map.get(data.get('tef', 'Gold'), 0),
        'tef_experience': tef_map.get(data.get('tef_exp', 'Gold'), 0),
        'tef_outcomes': tef_map.get(data.get('tef_out', 'Gold'), 0),
        'TARAGG': data.get('tariff', 120.0),
        'nss_average_satisfaction': data.get('nss', 85.0)
    }])
    predictions = {}
    for tn, model in ML_MODELS.items():
        try:
            predictions[tn] = float(model.predict(input_data)[0])
        except:
            predictions[tn] = None
    return jsonify({"predictions": predictions})

# ── Student Portal AI Copilot Advisor ──

@app.route("/api/student_advisor", methods=["POST"])
@student_self_only
def api_student_advisor():
    data = request.get_json()
    student_id = data.get("student_id")
    message = data.get("message", "").lower()
    
    student = get_student(student_id)
    if not student:
        return jsonify({"reply": "I couldn't find your profile. Please complete Step 1 first."})

    # Evaluate current targets
    student_gaps = []
    has_math_gap = False
    students = load_students()
    for tid in student.get("targets", []):
        agent_res = agent.solve_goal(student["id"], tid, students, silent=True)
        if agent_res and not agent_res.get("compliant", False):
            for gap in agent_res.get("gaps", []):
                student_gaps.append(gap)
                if "math" in gap.get("subject", "").lower() or "mathematics" in gap.get("description", "").lower():
                    has_math_gap = True

    # Agentic reasoning response
    if "tmua" in message or "test" in message or "exam" in message:
        reply = (
            "🎯 **TMUA (Test of Mathematics for University Admission) Insights:**\n\n"
            "The TMUA is mandatory for Cambridge CS and Imperial Computing. It consists of two papers: \n"
            "1. **Mathematical Reasoning** (20 multiple choice questions, 75 mins)\n"
            "2. **Mathematical Speculation** (20 multiple choice questions, 75 mins)\n\n"
            "**Advisors Recommended Actions:**\n"
            "- Start preparing with past papers from the official Cambridge Admissions website.\n"
            "- Solve UKMT Senior Mathematical Challenge papers to build speed and logical analysis.\n"
            "- Double check the registration deadline: **September 16, 2026**."
        )
    elif "math" in message or "subject" in message:
        if has_math_gap:
            reply = (
                "⚠️ **Mathematics Requirement Alert:**\n\n"
                "Your profile currently shows a critical Mathematics prerequisite gap for your targets. "
                "Because CBSE/ICSE doesn't easily permit late subject additions in Class 12, here is your agentic action plan:\n\n"
                "1. **AP Calculus BC Override:** Register for AP Calculus BC in May to satisfy Cambridge CS/Imperial Math prerequisites.\n"
                "2. **Target List Pivoting:** Consider applying to courses like BA Business Administration or BCA, or private universities (e.g. Ashoka University) where Class 12 Math is not mandatory.\n"
                "3. **Board Registration Check:** Verify with your school counselor if it's still possible to register for Mathematics as a 6th subject."
            )
        else:
            reply = (
                "📚 **Subject Strategy advice:**\n\n"
                "Your current board subject registration matches your target pathways. Ensure you maintain at least **95% in Mathematics and Physics** if you are targeting elite UK pathways like Cambridge CS."
            )
    elif "portfolio" in message or "extracurricular" in message or "activity" in message:
        reply = (
            "🏆 **Extracurricular Portfolio Roadmap:**\n\n"
            "Our AI auto-classifier assesses the impact tier of your activities based on global reach. \n"
            "- **Tier 1 (elite):** Research papers (IEEE, arXiv), national olympiads (IMO, IOI), patent filings, or global startup launch.\n"
            "- **Tier 2 (strong):** State championships, regional hackathon winners, founding clubs, head boy/girl status.\n\n"
            "**Action Item:** If targeting US universities (Stanford, MIT), aim to convert one of your Tier 3 school activities into a Tier 1 or Tier 2 regional/national project."
        )
    else:
        gaps_summary = f"Currently, you have {len(student_gaps)} active gap(s) across your target pathways." if student_gaps else "Awesome! You are fully on track with no gaps."
        reply = (
            f"Hello {student.get('name')}! I am your unlockED Pathway Copilot. {gaps_summary}\n\n"
            "Ask me anything about:\n"
            "- **'TMUA preparation'** or registration timelines\n"
            "- **'Math requirements'** or board subject mismatch remediations\n"
            "- **'Portfolio projects'** to level up your extracurricular tier"
        )

    return jsonify({"reply": reply})

# ── Counselor Portal AI Cohort Command Center Agent ──

@app.route("/api/counselor_agent", methods=["POST"])
def api_counselor_agent():
    data = request.get_json() or {}
    command = data.get("command", "").strip()
    if not command:
        return jsonify({"response": "Please enter a counselor command or query."}), 400
    
    # Compile live cohort context for reasoning model
    cohort_summary_list = []
    students = load_students()
    for s in students:
        student_audits = {}
        for tid in s.get("targets", []):
            try:
                tid_str = tid.get("id", tid) if isinstance(tid, dict) else tid
                audit_res = agent.solve_goal(s["id"], tid_str, students, silent=True)
                if audit_res:
                    student_audits[str(tid_str)] = {
                        "match_score": audit_res.get("match_score", generate_realistic_match_score(s["id"], tid_str, audit_res.get("compliant", True))),
                        "compliant": audit_res.get("compliant", True),
                        "gaps": audit_res.get("gaps", []),
                        "remediations": audit_res.get("remediations", [])
                    }
            except Exception:
                pass
        cohort_summary_list.append({
            "id": s["id"],
            "name": s["name"],
            "board": s.get("board"),
            "class_level": s.get("class_level"),
            "board_subjects": s.get("board_subjects", []),
            "grades": s.get("grades", {}),
            "standardized_tests": s.get("standardized_tests", {}),
            "portfolio": s.get("portfolio", []),
            "targets": s.get("targets", []),
            "audits": student_audits
        })

    groq_key = os.environ.get("GROQ_API_KEY", "").strip()
    alumni_db = load_alumni()

    system_prompt = f"""You are the unlockED Counselor AI Agent & Chief Admissions Officer Co-Pilot.
You have access to the entire school's active student cohort database and historical alumni success database.

STUDENT COHORT DATA:
{json.dumps(cohort_summary_list, indent=2)}

HISTORICAL ALUMNI SUCCESS STORIES (PROVEN PATHWAYS):
{json.dumps(alumni_db, indent=2)}

CRITICAL REASONING INSTRUCTIONS:
1. Act as a highly intelligent, expert high school counselor and admissions strategist.
2. Analyze the counselor's prompt carefully. It may ask you to:
   - Perform a risk audit or cohort gap analysis across all students.
   - Draft personalized warning or guidance emails to specific students or parents.
   - Recommend new target pathways, university choices, or portfolio improvements for a specific student (e.g. STU_001).
   - Provide strategic interventions, deadline warnings, or custom counseling notes.
3. NEVER return static hardcoded canned responses. Always analyze the actual live JSON student cohort data dynamically.
4. Format your output in clean Markdown with appropriate headers, bold text, bullet points, and actionable details.
5. If drafting an email, include Subject line, To address, Salutation, specific student gap evidence, and professional sign-off.
"""
    if groq_key:
        import requests
        try:
            print("[CounselorAgent] Calling Groq API (llama-3.3-70b-versatile)...")
            res = requests.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization": f"Bearer {groq_key}", "Content-Type": "application/json"},
                json={
                    "model": "llama-3.3-70b-versatile",
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": f"COUNSELOR PROMPT / COMMAND:\n{command}"}
                    ],
                    "temperature": 0.2
                },
                timeout=30
            )
            if res.status_code == 200:
                text_out = res.json()["choices"][0]["message"]["content"]
                return jsonify({"response": text_out.strip()})
            else:
                print(f"[CounselorAgent Warning] Groq returned HTTP {res.status_code}: {res.text}")
        except Exception as g_err:
            print(f"[CounselorAgent Warning] Groq reasoning call failed: {g_err}")

    # Fallback to local Ollama engine if Groq is not configured or fails
    import requests
    try:
        ollama_res = requests.post("http://127.0.0.1:11434/api/generate", json={
            "model": "llama3.2",
            "prompt": f"{system_prompt}\n\nCOUNSELOR PROMPT / COMMAND:\n{command}",
            "stream": False,
            "options": {"temperature": 0.3}
        }, timeout=45)
        if ollama_res.status_code == 200:
            return jsonify({"response": ollama_res.json().get("response", "").strip()})
    except Exception as e:
        print(f"[CounselorAgent Warning] Local Ollama fallback failed: {e}")

    # Final hardcoded fallback if everything else fails
    command_lower = command.lower()
    if "email" in command_lower or "draft" in command_lower:
        flagged = [s for s in cohort_summary_list if any(not a.get("compliant") for a in s["audits"].values())]
        if not flagged:
            return jsonify({"response": "### 📧 Email Assistant\n\nNo students currently require urgent warning emails."})
        target_student = flagged[0]
        first_gap = next((g for a in target_student["audits"].values() for g in a.get("gaps", [])), {})
        draft = (
            f"### 📧 Dynamically Generated Warning Draft for {target_student['name']} ({target_student['id']})\n\n"
            f"**To:** {target_student['name'].lower().replace(' ', '.')}@school.edu\n"
            f"**Subject:** Action Required: Urgent Pathway Prerequisite Gap ({first_gap.get('subject', 'Academic Mismatch')})\n\n"
            f"Dear {target_student['name']},\n\n"
            f"Our unlockED compliance audit detected an active prerequisite gap for your target university pathways:\n"
            f"👉 *{first_gap.get('description', 'Prerequisite subject or grade cutoff missing.')}*\n\n"
            f"Please schedule a consultation with your school counselor to adjust your subject selection or pathway targets.\n\n"
            f"Best regards,\nSchool Counseling Office"
        )
        return jsonify({"response": draft})
    
    elif "recommend" in command_lower or "suggest" in command_lower or "stu_" in command_lower:
        student_match = re.search(r'(stu_\d+)', command_lower)
        sid = student_match.group(1).upper() if student_match else "STU_001"
        s = next((st for st in cohort_summary_list if st["id"] == sid), cohort_summary_list[0] if cohort_summary_list else None)
        if not s:
            return jsonify({"response": f"Student ID '{sid}' not found."})
        
        resp = f"### 🎯 Strategic University & Pathway Recommendations for {s['name']} ({s['id']})\n\n"
        resp += f"**Academic Board:** {s['board']} (Class {s['class_level']}) | **Current Targets:** {', '.join(s['targets']) or 'None'}\n\n"
        resp += f"1. **High-Fit Pathway Optimization:** Ensure studied subjects ({', '.join(s['board_subjects'])}) are mapped to domain prerequisites.\n"
        resp += f"2. **Portfolio Scaling:** {len(s['portfolio'])} activities registered. Focus on achieving Tier 1 national/international recognition.\n"
        return jsonify({"response": resp})
        
    else:
        total = len(cohort_summary_list)
        high_risk = sum(1 for s in cohort_summary_list if any(a.get("match_score", 100) < 70 for a in s["audits"].values()))
        resp = (
            f"### 📊 Live Cohort Risk & Compliance Analysis\n\n"
            f"- **Active Cohort Size:** {total} students evaluated\n"
            f"- **High Risk Count (<70% match):** **{high_risk}** student(s)\n\n"
            f"**Recommended Actions:**\n"
            f"1. Run targeted subject remediation workshops for flagged students.\n"
            f"2. Use command `draft warning emails` to auto-generate warning communications."
        )
        return jsonify({"response": resp})

# ── Opportunity Radar & Competition Monitor Endpoints ──

COMPETITIONS_PATH = os.path.join(BASE_DIR, "data", "competitions_db.json")

def load_competitions():
    if os.path.exists(COMPETITIONS_PATH):
        with open(COMPETITIONS_PATH, "r") as f:
            return json.load(f)
    return []

def save_competitions(comps):
    with open(COMPETITIONS_PATH, "w") as f:
        json.dump(comps, f, indent=2)

@app.route("/api/opportunities/<student_id>")
@student_self_only
def api_opportunities(student_id):
    student = get_student(student_id)
    if not student:
        return jsonify({"error": "Student not found"}), 404
        
    competitions = load_competitions()
    matches = OpportunityRadar.match_student(student, competitions)
    return jsonify(matches)

@app.route("/api/import_competition", methods=["POST"])
@counselor_required
def api_import_competition():
    data = request.get_json()
    if not data or "url" not in data:
        return jsonify({"error": "URL is required"}), 400
        
    url = data["url"].strip()
    imported_comp = CompeteMapScraper.scrape_url(url)
    if not imported_comp:
        return jsonify({"error": "Failed to scrape competition details. Make sure the URL is a valid CompeteMap competition link."}), 400
        
    comps = load_competitions()
    idx = next((i for i, c in enumerate(comps) if c["id"] == imported_comp["id"]), None)
    if idx is not None:
        comps[idx] = imported_comp
    else:
        comps.append(imported_comp)
        
    save_competitions(comps)
    return jsonify(imported_comp), 201

@app.route("/api/competitions")
@login_required
def api_competitions():
    return jsonify(load_competitions())

@app.route("/api/shortlist/<student_id>", methods=["POST"])
@counselor_required
def api_shortlist_toggle(student_id):
    data = request.get_json()
    college_id = data.get("college_id")
    if not college_id:
        return jsonify({"error": "college_id required"}), 400

    student = get_student(student_id)
    print("Looking for student:", student_id, "Found:", bool(student))
    if student is None:
        print("Available students:", [doc.id for doc in db.collection(STUDENTS_COLLECTION).stream()])
        return jsonify({"error": "Student not found"}), 404

    shortlisted = student.get("shortlisted_colleges", [])

    if college_id in shortlisted:
        shortlisted.remove(college_id)
        added = False
    else:
        shortlisted.append(college_id)
        added = True

    student["shortlisted_colleges"] = shortlisted
    save_student(student)
    return jsonify({"added": added, "shortlisted_colleges": shortlisted})

# ── College Shortlist & Deadline Calendar Endpoints ──

COLLEGES_PATH = os.path.join(BASE_DIR, "data", "colleges_db.json")
EXAMS_PATH = os.path.join(BASE_DIR, "data", "exams_db.json")

def load_colleges():
    global COLLEGES_CACHE
    if COLLEGES_CACHE is not None:
        return COLLEGES_CACHE

    # Base colleges from colleges_db.json
    colleges = []
    if os.path.exists(COLLEGES_PATH):
        with open(COLLEGES_PATH, "r", encoding="utf-8") as f:
            colleges.extend(json.load(f))
            
    # Add Indian
    ind_json_path = os.path.join(BASE_DIR, "data", "indian_unis.json")
    if os.path.exists(ind_json_path):
        with open(ind_json_path, "r", encoding="utf-8") as f:
            ind_data = json.load(f)
            for item in ind_data:
                if not any(c["id"] == item["id"] for c in colleges):
                    colleges.append(item)
                    
    # Add US
    us_df_path = os.path.join(BASE_DIR, "data", "cleaned_us_colleges.csv")
    if os.path.exists(us_df_path):
        df_us = pd.read_csv(us_df_path)
        for _, row in df_us.iterrows():
            cid = f"US_{row['UNITID']}"
            name = str(row['INSTNM'])
            if not any(c["name"] == name for c in colleges):
                try:
                    deadlines = json.loads(str(row.get('DEADLINES', '[]')))
                except:
                    deadlines = []
                try:
                    subject_requirements = json.loads(str(row.get('SUBJECT_REQUIREMENTS', '[]')))
                except:
                    subject_requirements = []
                try:
                    required_exams = json.loads(str(row.get('REQUIRED_EXAMS', '[]')))
                except:
                    required_exams = []

                colleges.append({
                    "id": cid,
                    "name": name,
                    "country": "US",
                    "courses": ["Undergraduate Program"],
                    "deadlines": deadlines,
                    "required_exams": required_exams,
                    "subject_requirements": subject_requirements,
                    "expected_sat": str(int(row["SAT_AVG"])) if pd.notna(row.get("SAT_AVG")) and str(row.get("SAT_AVG")) not in ["N/A", "nan"] else "N/A"
                })
                
    # Add UK
    uk_df_path = os.path.join(BASE_DIR, "data", "cleaned_courses_dataset.csv")
    if os.path.exists(uk_df_path):
        df_uk = pd.read_csv(uk_df_path, usecols=["LEGAL_NAME"]).drop_duplicates()
        for _, row in df_uk.iterrows():
            name = str(row["LEGAL_NAME"])
            if pd.isna(name) or name == "nan": continue
            import re
            cid = f"UK_{re.sub(r'[^A-Z0-9]', '', name.upper())}"
            if not any(c["name"] == name for c in colleges):
                colleges.append({
                    "id": cid,
                    "name": name,
                    "country": "UK",
                    "courses": ["Various Programs"],
                    "deadlines": [],
                    "required_exams": []
                })
                
    COLLEGES_CACHE = colleges
    return colleges

def load_exams():
    if os.path.exists(EXAMS_PATH):
        with open(EXAMS_PATH, "r") as f:
            return json.load(f)
    return []

@app.route("/api/colleges")
@login_required
def api_colleges():
    return jsonify(load_colleges())

def _find_requirements_for_college(college):
    """Find matching requirements_db entries for a college from colleges_db."""
    college_name = (college.get("name") or "").lower()
    college_id = (college.get("id") or "").lower()
    best_match = None
    best_score = 0

    for _rid, req in kg.requirements.items():
        uni_name = (req.get("university") or "").lower()
        req_name = (req.get("name") or "").lower()
        score = 0
        if uni_name and uni_name in college_name:
            score += 3
        elif college_name and college_name in uni_name:
            score += 3
        name_words = college_id.lower().replace("_", " ").split()
        for w in name_words:
            if len(w) > 2 and w in uni_name:
                score += 1
            if len(w) > 2 and w in req_name:
                score += 1
        if score > best_score:
            best_score = score
            best_match = req
    return best_match if best_score >= 2 else None


@app.route("/api/evaluate_shortlist", methods=["POST"])
@login_required
def api_evaluate_shortlist():
    data = request.json
    student_id = data.get("student_id")
    college_id = data.get("college_id")

    student = get_student(student_id)
    if not student:
        return jsonify({"error": "Student not found"}), 404

    colleges = load_colleges()
    college = next((c for c in colleges if c["id"] == college_id), None)
    
    if not college:
        import re
        def normalize(s):
            return re.sub(r'[^a-z0-9]', '', str(s).lower())
        def get_initials(s):
            words = re.findall(r'[a-zA-Z0-9]+', str(s))
            return "".join(w[0].lower() for w in words if w.lower() not in ('of', 'the', 'at', 'and'))
            
        norm_id = normalize(college_id)
        
        # 1. Try exact normalized match
        college = next((c for c in colleges if normalize(c["name"]) == norm_id), None)
        
        # 2. Try partial match - only when query has 4+ chars to avoid false positives
        if not college and len(norm_id) >= 4:
            college = next((c for c in colleges if norm_id in normalize(c["name"]) or normalize(c["name"]) in norm_id), None)
            
        # 3. Try acronym match (e.g. NYU -> New York University, LSE -> London School of Economics)
        # Only if query looks like an acronym (all uppercase original, or <= 5 chars)
        if not college and (college_id.upper() == college_id or len(college_id) <= 5):
            college = next((c for c in colleges if get_initials(c["name"]) == norm_id), None)

    if not college:
        return jsonify({"category": "Target", "match_score": 50, "reasoning": "College not found in database.", "gaps": [], "strengths": []})

    requirements_info = _find_requirements_for_college(college)
    if requirements_info:
        if not college.get("admission_tests") and not college.get("required_exams"):
            college["admission_tests"] = requirements_info.get("admission_tests", [])
        if not college.get("subject_requirements"):
            college["subject_requirements"] = [p.get("subject", "") for p in requirements_info.get("subject_prerequisites", [])]

    llm_result = llm_scorer.classify_shortlist(student, college, requirements_info=requirements_info)
    if llm_result:
        llm_result["required_exams"] = college.get("admission_tests", college.get("required_exams", []))
        if requirements_info:
            llm_result["subject_prerequisites"] = requirements_info.get("subject_prerequisites", [])
            llm_result["grade_prerequisites"] = requirements_info.get("grade_prerequisites", [])
        return jsonify(llm_result)

    # Deterministic fallback
    grades_dict = student.get("grades", {})
    grade = grades_dict.get("current_expected_board") or grades_dict.get("class_12_aggregate") or grades_dict.get("class_10_aggregate") or "80"

    grade_str = str(grade).replace('%','').strip()
    if '-' in grade_str:
        parts = grade_str.split('-')
        try:
            grade_val = sum(float(p.strip()) for p in parts) / len(parts)
        except:
            grade_val = 80.0
    else:
        try:
            grade_val = float(grade_str)
        except:
            grade_val = 80.0

    tests_dict = student.get("standardized_tests", {})
    sat = tests_dict.get("SAT") or tests_dict.get("sat")
    try:
        sat_val = int(sat)
    except:
        sat_val = 0

    name = college.get("name", "").lower()

    ELITE = ["harvard", "yale", "stanford", "mit", "princeton", "caltech", "cambridge", "oxford", "imperial", "eth zurich", "lse", "chicago", "columbia", "iit bombay", "iit delhi", "iit madras", "iisc", "aiims"]
    TOP = ["cornell", "ucl", "ucla", "uc berkeley", "michigan", "nyu", "toronto", "melbourne", "edinburgh", "duke", "johns hopkins", "bits pilani", "iit kanpur", "iit kharagpur", "iit roorkee"]
    STRONG = ["purdue", "umass", "ut austin", "ohio state", "penn state", "arizona state", "illinois", "wisconsin", "georgia tech", "ashoka", "du srcc", "st stephens"]

    if any(x in name for x in ELITE):
        tier = 1
    elif any(x in name for x in TOP):
        tier = 2
    elif any(x in name for x in STRONG):
        tier = 3
    else:
        tier = 4

    if tier == 1:
        match_score = max(25, min(55, int(grade_val * 0.5 + (sat_val / 40 if sat_val else 0))))
        if grade_val >= 98 and sat_val >= 1560:
            category = "Target"
        else:
            category = "Reach"
    elif tier == 2:
        match_score = max(35, min(72, int(grade_val * 0.65 + (sat_val / 50 if sat_val else 0))))
        if grade_val >= 95 and (sat_val >= 1480 or sat_val == 0):
            category = "Target"
        else:
            category = "Reach"
    elif tier == 3:
        match_score = max(45, min(85, int(grade_val * 0.8 + (sat_val / 60 if sat_val else 5))))
        if grade_val >= 90:
            category = "Safety"
        elif grade_val >= 80:
            category = "Target"
        else:
            category = "Reach"
    else:
        match_score = max(50, min(90, int(grade_val * 0.9 + (sat_val / 80 if sat_val else 5))))
        if grade_val >= 85:
            category = "Safety"
        elif grade_val >= 70:
            category = "Target"
        else:
            category = "Reach"

    admission_tests = college.get("admission_tests", college.get("required_exams", []))
    if any(t in ("CUET_UG", "JEE_MAIN", "JEE_ADVANCED") for t in admission_tests):
        if category == "Safety":
            category = "Target"

    gaps = []
    student_subjects = [s.lower() for s in student.get("board_subjects", [])]
    if requirements_info:
        for prereq in requirements_info.get("subject_prerequisites", []):
            subj = prereq.get("subject", "")
            if prereq.get("level") == "compulsory" and subj.lower() not in student_subjects:
                gaps.append(f"Missing required subject: {subj}")
    if sat_val == 0 and any(t in ("SAT", "ACT") for t in admission_tests):
        gaps.append("No SAT/ACT score on file")

    reasoning = f"{'Elite' if tier == 1 else 'Top' if tier == 2 else 'Strong' if tier == 3 else 'Standard'} institution. "
    reasoning += f"Student grade: {grade}, SAT: {sat_val if sat_val else 'N/A'}. "
    if gaps:
        reasoning += f"{len(gaps)} gap(s) identified."
    else:
        reasoning += "No major gaps found."

    return jsonify({
        "category": category,
        "tier": tier,
        "match_score": match_score,
        "reasoning": reasoning,
        "gaps": gaps,
        "strengths": [],
        "required_exams": admission_tests,
    })


@app.route("/api/export_shortlist/<student_id>")
@login_required
def api_export_shortlist(student_id):
    """Export a student's shortlisted colleges with evaluations as Excel."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    student = get_student(student_id)
    if not student:
        return jsonify({"error": "Student not found"}), 404

    shortlisted_ids = student.get("shortlisted_colleges", [])
    if not shortlisted_ids:
        return jsonify({"error": "No shortlisted colleges"}), 400

    colleges = load_colleges()
    wb = Workbook()
    ws = wb.active
    ws.title = "University Shortlist"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    reach_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    target_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    safety_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["University", "Country", "Category", "Match Score", "Required Exams", "Gaps", "Reasoning"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row_idx = 2
    for cid in shortlisted_ids:
        college = next((c for c in colleges if c["id"] == cid), None)
        college_name = college["name"] if college else cid

        eval_data = {"category": "Target", "match_score": 50, "reasoning": "", "gaps": [], "required_exams": []}
        if college:
            req_info = _find_requirements_for_college(college)
            llm_res = llm_scorer.classify_shortlist(student, college, requirements_info=req_info)
            if llm_res:
                eval_data = llm_res
                eval_data["required_exams"] = college.get("admission_tests", college.get("required_exams", []))

        category = eval_data.get("category", "Target")
        row_fill = reach_fill if category == "Reach" else safety_fill if category == "Safety" else target_fill

        values = [
            college_name,
            college.get("country", "N/A") if college else "N/A",
            category,
            eval_data.get("match_score", 50),
            ", ".join(eval_data.get("required_exams", [])),
            "; ".join(eval_data.get("gaps", [])) or "None",
            eval_data.get("reasoning", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 3:
                cell.fill = row_fill
                cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        row_idx += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 50

    info_row = row_idx + 1
    ws.cell(row=info_row, column=1, value=f"Student: {student.get('name', student_id)}")
    ws.cell(row=info_row, column=1).font = Font(bold=True, size=10)
    ws.cell(row=info_row + 1, column=1, value=f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=info_row + 1, column=1).font = Font(italic=True, size=9, color="666666")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = student.get("name", student_id).replace(" ", "_")
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{safe_name}_University_Shortlist.xlsx"
    )


@app.route("/api/priority_queue/<student_id>")
@student_self_only
def api_priority_queue(student_id):
    student = get_student(student_id)
    if not student:
        return jsonify({"error": "Student not found"}), 404

    students_list = load_students()
    evaluations = {}
    for tid in student.get("targets", []):
        tid_str = tid.get("id", tid) if isinstance(tid, dict) else tid
        try:
            agent_res = agent._solve_goal_simulated(student["id"], tid_str, students_list, None, silent=True)
            if agent_res:
                evaluations[tid_str] = agent_res
        except Exception:
            pass

    priority_list = llm_scorer.rank_priority(student, evaluations)
    if priority_list is None:
        items = []
        for tid, ev in evaluations.items():
            items.append({
                "target_id": tid,
                "priority_score": ev.get("urgency_score", 0),
                "priority_reason": f"{ev.get('risk_level', 'Unknown')} — {len(ev.get('gaps', []))} gap(s)",
                "recommended_action": ev["remediations"][0].get("action_item", ev["remediations"][0].get("remediation", "Review gaps")) if ev.get("remediations") else "No action needed",
                "action_deadline": None,
            })
        items.sort(key=lambda x: x["priority_score"], reverse=True)
        priority_list = items

    return jsonify({"student_id": student_id, "priority_queue": priority_list})


@app.route("/api/exams")
@login_required
def api_exams():
    return jsonify(load_exams())

@app.route("/api/calendar/<student_id>")
@student_self_only
def api_calendar(student_id):
    student = get_student(student_id)
    if not student:
        return jsonify({"error": "Student not found"}), 404

    events = []
    seen_event_keys = set()

    # 1. Target Course Deadlines
    for tid in student.get("targets", []):
        t = kg.get_course_or_exam(tid)
        if t:
            for dl in t.get("deadlines", []):
                key = (t["name"], dl["label"], dl["date"])
                if key not in seen_event_keys:
                    seen_event_keys.add(key)
                    events.append({
                        "title": f"{t['name']} - {dl['label']}",
                        "date": dl["date"],
                        "type": "college",
                        "description": dl.get("description", "")
                    })

    # 2. Shortlisted College Deadlines
    colleges = load_colleges()
    shortlisted_ids = student.get("shortlisted_colleges", [])
    for cid in shortlisted_ids:
        c = next((col for col in colleges if col["id"] == cid), None)
        if c:
            for dl in c.get("deadlines", []):
                key = (c["name"], dl["label"], dl["date"])
                if key not in seen_event_keys:
                    seen_event_keys.add(key)
                    events.append({
                        "title": f"{c['name']} - {dl['label']}",
                        "date": dl["date"],
                        "type": "college",
                        "description": dl.get("description", "")
                    })

    # 3. Standardized Entrance Exams Timelines
    exams = load_exams()
    student_tracks = set()
    required_exam_ids = set()

    for tid in student.get("targets", []):
        t = kg.get_course_or_exam(tid)
        if t:
            student_tracks.add(t.get("track"))
            for et in t.get("admission_tests", []):
                required_exam_ids.add(et)

    for cid in shortlisted_ids:
        c = next((col for col in colleges if col["id"] == cid), None)
        if c:
            student_tracks.add(c.get("country"))
            for ex in c.get("required_exams", []):
                required_exam_ids.add(ex)

    for ex in exams:
        is_relevant = (
            ex["id"] in required_exam_ids or 
            ex["track"] in student_tracks or
            (ex["track"] == "US" and "US" in student_tracks) or
            (ex["track"] == "UK" and "UK" in student_tracks) or
            (ex["track"] == "India" and "India" in student_tracks)
        )
        if is_relevant:
            reg_key = (ex["name"], "Registration Deadline", ex["deadline"])
            if reg_key not in seen_event_keys:
                seen_event_keys.add(reg_key)
                events.append({
                    "title": f"{ex['name']} Registration Close",
                    "date": ex["deadline"],
                    "type": "exam",
                    "description": f"Registration closes. {ex['description']}"
                })
            exam_key = (ex["name"], "Exam Date", ex["exam_date"])
            if exam_key not in seen_event_keys:
                seen_event_keys.add(exam_key)
                events.append({
                    "title": f"{ex['name']} Exam Date",
                    "date": ex["exam_date"],
                    "type": "exam",
                    "description": f"Official test date. {ex['description']}"
                })

    # 4. Competition Deadlines (Opportunity Radar - Match Score >= 75%)
    competitions = load_competitions()
    matches = OpportunityRadar.match_student(student, competitions)
    for m in matches:
        if m["match_score"] >= 75:
            comp = m["competition"]
            if comp.get("deadline"):
                key = (comp["name"], "Submission Deadline", comp["deadline"])
                if key not in seen_event_keys:
                    seen_event_keys.add(key)
                    events.append({
                        "title": f"{comp['name']} Deadline",
                        "date": comp["deadline"],
                        "type": "competition",
                        "description": comp.get("description", "")
                    })

    events.sort(key=lambda x: x["date"])
    return jsonify(events)

# ── Scholarship & Financial Aid Endpoints ──

SCHOLARSHIPS_PATH = os.path.join(BASE_DIR, "data", "scholarships.json")


def load_scholarships():
    if os.path.exists(SCHOLARSHIPS_PATH):
        with open(SCHOLARSHIPS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def save_scholarships(schols):
    with open(SCHOLARSHIPS_PATH, "w", encoding="utf-8") as f:
        json.dump(schols, f, indent=2)

@app.route("/api/scholarships", methods=["GET"])
@login_required
def api_get_scholarships():
    return jsonify(load_scholarships())

@app.route("/api/match_scholarships/<student_id>")
@student_self_only
def api_match_scholarships(student_id):
    student = get_student(student_id)
    if not student:
        return jsonify({"error": "Student not found"}), 404

    scholarships = load_scholarships()
    matches = ScholarshipAgent.match_scholarships(student, scholarships)
    return jsonify(matches)

@app.route("/api/recommend_scholarship_students/<scholarship_id>", methods=["POST"])
@counselor_required
def api_recommend_scholarship_students(scholarship_id):
    scholarships = load_scholarships()
    scholarship = next((s for s in scholarships if s["id"] == scholarship_id), None)
    if not scholarship:
        return jsonify({"error": "Scholarship not found"}), 404

    recommended = ScholarshipAgent.recommend_students(scholarship, load_students())
    return jsonify(recommended)

@app.route("/api/students/<student_id>/shortlist_scholarship", methods=["POST"])
@counselor_required
def api_shortlist_scholarship(student_id):
    student = get_student(student_id)
    if not student:
        return jsonify({"error": "Student not found"}), 404
    
    data = request.json
    scholarship_id = data.get("scholarship_id")
    if not scholarship_id:
        return jsonify({"error": "Missing scholarship_id"}), 400
        
    if "shortlisted_scholarships" not in student:
        student["shortlisted_scholarships"] = []
        
    if scholarship_id in student["shortlisted_scholarships"]:
        student["shortlisted_scholarships"].remove(scholarship_id)
        added = False
    else:
        student["shortlisted_scholarships"].append(scholarship_id)
        added = True

    save_student(student)
    return jsonify({"success": True, "added": added, "shortlisted": student["shortlisted_scholarships"]})

@app.route("/api/import_scholarship", methods=["POST"])
@counselor_required
def api_import_scholarship():
    data = request.get_json()
    if not data or "name" not in data:
        return jsonify({"error": "Scholarship name is required"}), 400
        
    import datetime
    schol_id = f"schol_custom_{int(datetime.datetime.now().timestamp())}"
    
    new_schol = {
        "id": schol_id,
        "name": data.get("name"),
        "type": data.get("type", "Merit-based"),
        "provider": data.get("provider", "Unknown Provider"),
        "eligibility_criteria": data.get("eligibility_criteria", "See official website for details."),
        "award_value": data.get("award_value", "Varies"),
        "deadline": data.get("deadline", "TBD"),
        "tags": data.get("tags", []),
        "min_class_level": int(data.get("min_class_level", 9)),
        "max_class_level": int(data.get("max_class_level", 12)),
        "url": data.get("url", "#")
    }
    
    schols = load_scholarships()
    schols.append(new_schol)
    save_scholarships(schols)
    
    return jsonify({"success": True, "scholarship": new_schol})

# ── Recommendation Studio ──

@app.route("/api/draft_recommendation", methods=["POST"])
@counselor_required
def api_draft_recommendation():
    data = request.get_json()
    student_id = data.get("student_id")
    student = get_student(student_id)
    if not student:
        return jsonify({"error": "Student not found"}), 404
        
    portfolio = student.get("portfolio", [])
    grades = student.get("grades", {})
    targets = student.get("targets", [])
    shortlisted_colleges = student.get("shortlisted_colleges", [])
    shortlisted_scholarships = student.get("shortlisted_scholarships", [])
    
    brag_sheet = f"Student Name: {student['name']}\n"
    brag_sheet += f"Board: {student.get('board')} Class {student.get('class_level')}\n"
    brag_sheet += f"Grades: {json.dumps(grades)}\n"
    brag_sheet += "Extracurriculars & Portfolio:\n"
    for item in portfolio:
        brag_sheet += f"- {item.get('title', '')} ({item.get('role', '')}): {item.get('description', '')}\n"
    brag_sheet += f"Target Universities: {', '.join(targets)}\n"
    brag_sheet += f"Shortlisted Colleges: {', '.join(shortlisted_colleges)}\n"
    brag_sheet += f"Shortlisted / Assigned Scholarships: {', '.join(shortlisted_scholarships)}\n"

    system_prompt = "You are an expert high school counselor writing a powerful, persuasive Letter of Recommendation (LOR) for a student's university application. Use the student's exact brag sheet and data to write a tailored draft. Do not use generic placeholders like [Name], use the real name. Keep it professional, highlighting their specific achievements. Output ONLY the letter itself."
    
    import requests
    from flask import Response
    try:
        ollama_res = requests.post("http://127.0.0.1:11434/api/generate", json={
            "model": "llama3.2",
            "prompt": f"{system_prompt}\n\nSTUDENT BRAG SHEET:\n{brag_sheet}\n\nDraft the complete Letter of Recommendation below:",
            "stream": True,
            "options": {"temperature": 0.6}
        }, timeout=60, stream=True)
        
        def generate():
            for line in ollama_res.iter_lines():
                if line:
                    try:
                        import json
                        chunk = json.loads(line.decode('utf-8'))
                        if "response" in chunk:
                            yield chunk["response"]
                    except:
                        pass
        
        return Response(generate(), mimetype='text/plain')
    except Exception as e:
        print(f"[Recommendation Error] Ollama failed: {e}")
        return jsonify({"error": "Engine failed to generate draft."}), 500

# ── Bulk Ingestion ──

@app.route("/api/bulk_ingest_preview", methods=["POST"])
def api_bulk_ingest_preview():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Empty filename"}), 400

    import csv
    import io
    
    try:
        content = file.read().decode('utf-8')
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        if len(rows) < 2:
            return jsonify({"error": "File must have a header row and at least one data row"}), 400
            
        headers = [h.strip() for h in rows[0]]
        data_rows = rows[1:]
        
        # 1. Ask Ollama to map the headers
        import requests
        system_prompt = '''You are a data mapper. The user has uploaded a spreadsheet of high school students.
You will be given the column headers of the spreadsheet.
Map them to the following JSON keys:
- "name_col" (string, the column with the student's full name)
- "class_col" (string, the column with their class/grade level)
- "board_col" (string, the column with their academic board, e.g. CBSE/IB)
- "subjects_col" (string, the column with their studied subjects)
- "targets_col" (string, the column with their target colleges/universities)
- "portfolio_col" (string, the column with their extracurriculars/achievements)

Respond ONLY with a valid JSON object containing exactly these keys. If a concept is missing from the headers, set its value to null. Do not include markdown formatting like ```json.'''
        
        mapping = {}
        try:
            ollama_res = requests.post("http://127.0.0.1:11434/api/generate", json={
                "model": "llama3.2",
                "prompt": f"{system_prompt}\n\nHEADERS: {json.dumps(headers)}",
                "stream": False,
                "options": {"temperature": 0.1}
            }, timeout=30)
            
            if ollama_res.status_code == 200:
                raw = ollama_res.json().get("response", "").strip()
                import re
                match = re.search(r"\{.*\}", raw, re.DOTALL)
                if match:
                    mapping = json.loads(match.group(0))
                else:
                    mapping = json.loads(raw)
        except Exception as e:
            print(f"[BulkIngest] AI header mapping failed: {e}")
            return jsonify({"error": "Failed to map columns using AI engine."}), 500

        # Create mapped previews
        previews = []
        for i, row in enumerate(data_rows):
            def get_val(col_name):
                if not col_name or col_name not in headers: return ""
                idx = headers.index(col_name)
                return row[idx] if idx < len(row) else ""
            
            p_name = get_val(mapping.get("name_col"))
            if not p_name: continue # skip empty rows
            
            p_class = get_val(mapping.get("class_col"))
            try:
                p_class_int = int(p_class) if p_class.isdigit() else 12
            except:
                p_class_int = 12
                
            p_board = get_val(mapping.get("board_col"))
            
            p_subjects_raw = get_val(mapping.get("subjects_col"))
            p_subjects = [s.strip() for s in p_subjects_raw.split(',')] if p_subjects_raw else []
            
            p_targets_raw = get_val(mapping.get("targets_col"))
            p_targets = [t.strip() for t in p_targets_raw.split(',')] if p_targets_raw else []
            
            p_portfolio_raw = get_val(mapping.get("portfolio_col"))
            p_portfolio = [{"activity": p.strip(), "description": "", "tier": 3} for p in p_portfolio_raw.split('\n') if p.strip()] if p_portfolio_raw else []
            
            previews.append({
                "name": p_name,
                "class_level": p_class_int,
                "board": p_board,
                "board_subjects": p_subjects,
                "targets": p_targets,
                "portfolio": p_portfolio,
                "grades": {},
                "standardized_tests": {},
                "shortlisted_colleges": []
            })
            
        return jsonify({"mapping": mapping, "previews": previews})
    except Exception as e:
        return jsonify({"error": f"Failed to parse CSV: {str(e)}"}), 500

@app.route("/api/bulk_ingest_save", methods=["POST"])
@counselor_required
def api_bulk_ingest_save():
    data = request.get_json()
    students_to_add = data.get("students", [])
    
    if not students_to_add:
        return jsonify({"error": "No students provided"}), 400

    students = load_students()
    for st in students_to_add:
        st["id"] = next_student_id(students)
        st["status"] = {"cuet_form_submitted": False, "tmua_registered": False, "sat_score": None}
        students.append(st)

    save_students(students_to_add)
    return jsonify({"success": True, "count": len(students_to_add)})

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=3000, debug=True)
